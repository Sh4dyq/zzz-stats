#!/usr/bin/env python3
"""Build a human-editable Excel of synergy tags from the sampled draft.

Model (per user 2026-07-14):
  - roles are MULTI-select (Miyabi = main crit carry AND main anomalist;
    Vivian = sub-anomalist only).
  - one shared TAG vocabulary; each tag can be checked under BOTH "gives" and
    "needs" (a carry that needs amp_on_stun; a stunner that gives it).
Rows = all 57 agents (names/element/specialty from skillsets_raw.json); the 6
sampled agents are pre-checked, the rest left blank for tagging.
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(__file__)
RAW = json.load(open(os.path.join(HERE, "skillsets_raw.json"), encoding="utf-8"))["agents"]

ROLES = [("crit_dps", "крит DPS"), ("sheer_dps", "sheer DPS (разрушение)"),
         ("sub_dps", "саб DPS"), ("main_anomaly", "мейн аномалист"),
         ("sub_anomaly", "саб аномалист"), ("stunner", "стан"),
         ("support", "саппорт"), ("off_field", "оф-филд")]
TAGS = [("atk_buff", "бафф ATK (ширикам ~0.5, скейлятся от HP)"),
        ("dmg_buff", "универс. DMG% / DMG-taken / RES-shred (полезен всем)"),
        ("crit_buff", "крит бафф (CRIT Rate/DMG)"),
        ("anomaly_buff", "бафф аномалий (Buildup/Proficiency/Disorder DMG)"),
        ("sheer_dmg_buff", "бафф sheer-урона / Sheer Force"),
        ("pen_buff", "бафф PEN команде"),
        ("def_shred", "снижение DEF врага"),
        ("daze", "стан/дазе"),
        ("amp_on_stun", "амп по застанненному (множитель + продление стана)"),
        ("anomaly_assist", "разгон аномалий/Disorder союзникам"),
        ("decibel", "децибелы (полезно, редко «нужно»)"),
        ("aftershock", "Aftershock (кол-во + бафы)"),
        ("abloom", "Abloom (для аномалистов: кол-во + бафы)"),
        ("ether_veil", "Ether Veil")]

# Full draft ratings, 0-5 scale. id -> (roles{}, gives{}, needs{}, note).
# Rated from Core Passive + Additional Ability (+ specialty/element/faction).
# Agents with no skill prose on nanoka yet are rated from role only + noted.
NODATA = "нет прозы скиллов на nanoka — размечено по амплуа, проверить"
RATINGS = {
 # ---- calibration six (top) ----
 "1191": ({"crit_dps": 5}, {}, {"daze": 5, "crit_buff": 5, "dmg_buff": 3, "atk_buff": 3}, ""),
 "1251": ({"stunner": 5}, {"daze": 5, "amp_on_stun": 3}, {"atk_buff": 3, "dmg_buff": 2}, ""),
 "1261": ({"main_anomaly": 5}, {"anomaly_assist": 3}, {"anomaly_assist": 4, "anomaly_buff": 4, "dmg_buff": 3},
          "может быть sub_anomaly, но только под Alice (парная связка, не общий кейс)"),
 "1311": ({"support": 5}, {"atk_buff": 5, "decibel": 4, "dmg_buff": 3}, {}, ""),
 "1091": ({"crit_dps": 5, "main_anomaly": 5}, {"anomaly_assist": 3},
          {"anomaly_assist": 4, "anomaly_buff": 3, "crit_buff": 4, "dmg_buff": 3, "atk_buff": 2}, ""),
 "1431": ({"sheer_dps": 5}, {}, {"amp_on_stun": 4, "ether_veil": 5, "sheer_dmg_buff": 4, "dmg_buff": 3, "atk_buff": 1}, ""),
 # ---- rest ----
 "1011": ({"stunner": 5}, {"daze": 4}, {"atk_buff": 2, "dmg_buff": 2}, ""),
 "1021": ({"crit_dps": 4, "sub_dps": 3}, {}, {"crit_buff": 3, "atk_buff": 3, "dmg_buff": 3, "anomaly_assist": 2}, ""),
 "1031": ({"support": 5}, {"def_shred": 4, "dmg_buff": 3}, {}, ""),
 "1041": ({"crit_dps": 4}, {}, {"crit_buff": 3, "atk_buff": 3, "dmg_buff": 3, "daze": 3}, ""),
 "1051": ({"sheer_dps": 5}, {}, {"sheer_dmg_buff": 3, "dmg_buff": 3, "ether_veil": 3, "atk_buff": 1}, "HP-скейл; AA-гейт: Stun/Support"),
 "1061": ({"crit_dps": 4}, {}, {"crit_buff": 3, "atk_buff": 3, "dmg_buff": 3, "daze": 4}, ""),
 "1071": ({"support": 4}, {"dmg_buff": 4}, {}, "щит + DMG-taken debuff 25%"),
 "1081": ({"crit_dps": 3, "sub_dps": 3}, {}, {"crit_buff": 3, "atk_buff": 3, "dmg_buff": 3}, ""),
 "1101": ({"stunner": 5}, {"daze": 4}, {"atk_buff": 2, "dmg_buff": 2}, ""),
 "1111": ({"crit_dps": 4}, {}, {"crit_buff": 3, "atk_buff": 3, "dmg_buff": 3, "anomaly_assist": 2}, ""),
 "1121": ({"support": 3}, {"crit_buff": 3}, {}, "щит + CRIT Rate +16%"),
 "1131": ({"support": 5}, {"atk_buff": 5, "dmg_buff": 3}, {}, ""),
 "1141": ({"stunner": 5}, {"daze": 4, "amp_on_stun": 4, "dmg_buff": 3}, {}, "Ice RES -25%"),
 "1151": ({"support": 5}, {"atk_buff": 4, "dmg_buff": 3}, {}, ""),
 "1161": ({"stunner": 5}, {"daze": 3, "dmg_buff": 3}, {}, ""),
 "1171": ({"main_anomaly": 5, "off_field": 4}, {"anomaly_assist": 3}, {"anomaly_assist": 3, "anomaly_buff": 3, "dmg_buff": 3}, ""),
 "1181": ({"main_anomaly": 5, "sub_anomaly": 3}, {"anomaly_assist": 2}, {"anomaly_buff": 3, "anomaly_assist": 3, "dmg_buff": 3}, ""),
 "1201": ({"crit_dps": 4}, {}, {"crit_buff": 3, "atk_buff": 3, "dmg_buff": 3, "daze": 4, "anomaly_assist": 3}, "AA: Stun/Anomaly"),
 "1211": ({"support": 5}, {"pen_buff": 5, "dmg_buff": 3}, {}, ""),
 "1221": ({"main_anomaly": 5}, {"anomaly_assist": 4}, {"anomaly_buff": 3, "anomaly_assist": 3, "dmg_buff": 3}, "Disorder mult +125%"),
 "1241": ({"crit_dps": 5}, {}, {"crit_buff": 4, "daze": 4, "dmg_buff": 3, "atk_buff": 3}, "AA-гейт: Support"),
 "1271": ({"support": 4}, {"dmg_buff": 3, "anomaly_assist": 3}, {}, "Anomaly Buildup RES -20%"),
 "1281": ({"main_anomaly": 5, "off_field": 3}, {"dmg_buff": 4}, {"anomaly_buff": 3, "anomaly_assist": 3, "dmg_buff": 3}, "squad DMG +18%"),
 "1291": ({"crit_dps": 5}, {}, {"crit_buff": 4, "daze": 5, "dmg_buff": 3, "atk_buff": 3}, "ATK масштаб от кол-ва Stun в отряде"),
 "1301": ({"crit_dps": 4, "off_field": 4}, {"aftershock": 3}, {"crit_buff": 3, "atk_buff": 3, "dmg_buff": 3, "daze": 3}, "AA: Stun/Support; aftershock DEF ignore"),
 "1321": ({"crit_dps": 5}, {}, {"crit_buff": 4, "atk_buff": 3, "dmg_buff": 3}, "AA-гейт: Stun/Support"),
 "1331": ({"sub_anomaly": 5, "off_field": 3}, {"anomaly_assist": 4}, {"anomaly_assist": 4, "anomaly_buff": 4, "dmg_buff": 3}, "ТОЛЬКО саб-аномалист (user)"),
 "1341": ({"support": 4}, {"dmg_buff": 4}, {"ether_veil": 4}, "squad DMG +10..40% внутри Ether Veil — нужен Ether Veil"),
 "1351": ({"stunner": 5}, {"daze": 3}, {"atk_buff": 2, "dmg_buff": 2}, "AA: Attack/Rupture"),
 "1361": ({"stunner": 5, "off_field": 3}, {"daze": 4, "amp_on_stun": 4, "aftershock": 4}, {"crit_buff": 2}, ""),
 "1371": ({"sheer_dps": 5, "main_anomaly": 3}, {}, {"sheer_dmg_buff": 4, "dmg_buff": 3, "atk_buff": 1, "ether_veil": 2}, "AA: Stun/Support/Defense; HP-скейл"),
 "1381": ({"crit_dps": 5}, {"aftershock": 4}, {"crit_buff": 3, "atk_buff": 3, "dmg_buff": 3, "aftershock": 3}, "AA: Stun/Support; squad Aftershock +25%"),
 "1391": ({"stunner": 5, "off_field": 3}, {"daze": 3, "decibel": 4}, {}, "max Decibel +1000"),
 "1401": ({"main_anomaly": 5}, {"anomaly_assist": 4}, {"anomaly_assist": 4, "anomaly_buff": 4, "dmg_buff": 3}, "связка с Jane"),
 "1411": ({"support": 5}, {"dmg_buff": 3}, {}, NODATA),
 "1421": ({"support": 4}, {"sheer_dmg_buff": 4, "dmg_buff": 4}, {}, "Sheer Force бафф + Depleted Qi +20%"),
 "1441": ({"sheer_dps": 5}, {}, {"sheer_dmg_buff": 4, "dmg_buff": 3, "atk_buff": 1, "ether_veil": 2}, "AA: Support/Stun; HP-скейл"),
 "1451": ({"support": 5}, {"ether_veil": 5, "dmg_buff": 3, "crit_buff": 3}, {}, "активирует Ether Veil: Wellspring — ключ для sheer"),
 "1461": ({"crit_dps": 3, "support": 3}, {"atk_buff": 3, "crit_buff": 3}, {}, "бафает Vanguard (Attack-напарника)"),
 "1471": ({"sheer_dps": 5}, {}, {"sheer_dmg_buff": 4, "dmg_buff": 3, "atk_buff": 1}, NODATA),
 "1481": ({"stunner": 5}, {"daze": 4}, {"atk_buff": 2, "dmg_buff": 2}, NODATA),
 "1491": ({"support": 5}, {"dmg_buff": 3, "ether_veil": 3}, {}, NODATA),
 "1501": ({"main_anomaly": 5}, {"anomaly_assist": 3}, {"anomaly_buff": 4, "anomaly_assist": 3, "dmg_buff": 3}, NODATA),
 "1511": ({"stunner": 5}, {"daze": 4}, {"atk_buff": 2, "dmg_buff": 2}, NODATA),
 "1521": ({"crit_dps": 4}, {}, {"crit_buff": 3, "atk_buff": 3, "dmg_buff": 3}, NODATA),
 "1531": ({"sheer_dps": 5}, {}, {"sheer_dmg_buff": 4, "dmg_buff": 3, "atk_buff": 1}, NODATA),
 "1541": ({"main_anomaly": 5}, {"anomaly_assist": 4}, {"anomaly_buff": 4, "anomaly_assist": 3, "dmg_buff": 3}, "Abloom DMG squad; frostbite extend"),
 "1551": ({"crit_dps": 4}, {"decibel": 3}, {"crit_buff": 3, "atk_buff": 3, "dmg_buff": 3, "daze": 3}, "AA: Stun/Support"),
 "1561": ({"main_anomaly": 5}, {"anomaly_assist": 3}, {"anomaly_buff": 4, "anomaly_assist": 3, "dmg_buff": 3}, NODATA),
 "1571": ({"stunner": 5}, {"daze": 4}, {"atk_buff": 2, "dmg_buff": 2}, NODATA),
 "1591": ({"crit_dps": 5}, {}, {"crit_buff": 3, "atk_buff": 3, "dmg_buff": 3, "daze": 3}, "AA: Support/Stun; Contamination"),
}
SAMPLE_ORDER = ["1191", "1251", "1261", "1311", "1091", "1431"]

# User edits are the source of truth: if a captured backup exists, use its
# 1-4 ratings verbatim (no rescaling) instead of the hardcoded RATINGS draft.
CUR = os.path.join(HERE, "synergy_tags_current.json")
if os.path.exists(CUR):
    _cur = json.load(open(CUR, encoding="utf-8"))
    RATINGS = {cid: (v["roles"], v["gives"], v["needs"], v.get("note", ""))
               for cid, v in _cur.items()}


def s3(v):
    return v or None


CHK = "✓"
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
H_ROLE = PatternFill("solid", fgColor="E8E4F5")
H_GIVE = PatternFill("solid", fgColor="DCEFE7")
H_NEED = PatternFill("solid", fgColor="FBEEDC")
H_META = PatternFill("solid", fgColor="F1EFE8")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center")


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Разметка"

    meta = ["агент", "элемент", "амплуа"]
    # column layout
    role_cols = [r[0] for r in ROLES]
    give_cols = [t[0] for t in TAGS]
    need_cols = [t[0] for t in TAGS]
    ncol = len(meta) + len(role_cols) + len(give_cols) + len(need_cols) + 1  # +заметки

    # row 1 = group banner, row 2 = column names
    c = 1
    for m in meta:
        ws.cell(2, c, m).font = Font(bold=True)
        ws.cell(2, c).fill = H_META
        c += 1
    grp_start = c
    for name in role_cols:
        ws.cell(2, c, name); ws.cell(2, c).fill = H_ROLE; c += 1
    ws.merge_cells(start_row=1, start_column=grp_start, end_row=1, end_column=c - 1)
    ws.cell(1, grp_start, "РОЛИ (несколько)").fill = H_ROLE

    grp_start = c
    for name in give_cols:
        ws.cell(2, c, name); ws.cell(2, c).fill = H_GIVE; c += 1
    ws.merge_cells(start_row=1, start_column=grp_start, end_row=1, end_column=c - 1)
    ws.cell(1, grp_start, "ДАЁТ команде").fill = H_GIVE

    grp_start = c
    for name in need_cols:
        ws.cell(2, c, name); ws.cell(2, c).fill = H_NEED; c += 1
    ws.merge_cells(start_row=1, start_column=grp_start, end_row=1, end_column=c - 1)
    ws.cell(1, grp_start, "НУЖНО от команды").fill = H_NEED

    notes_col = c
    ws.cell(2, notes_col, "заметки / условия").fill = H_META
    ws.cell(1, notes_col, "").fill = H_META

    hdr_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    for col in range(1, ncol + 1):
        ws.cell(1, col).font = Font(bold=True)
        ws.cell(1, col).alignment = center
        ws.cell(2, col).alignment = hdr_center
        ws.cell(2, col).font = Font(bold=True, size=10)

    # rows: calibration six first (for quick review), then the rest by id
    rest = [c for c in sorted(RAW, key=lambda x: int(x)) if c not in SAMPLE_ORDER]
    order = SAMPLE_ORDER + rest
    r = 3
    for cid in order:
        a = RAW[cid]
        ws.cell(r, 1, a["name"]).alignment = left
        ws.cell(r, 2, a.get("element") or "")
        ws.cell(r, 3, a.get("specialty") or "")
        base = len(meta)
        roles, gives, needs, note = RATINGS.get(cid, ({}, {}, {}, ""))
        for i, rc in enumerate(role_cols):
            if s3(roles.get(rc)): ws.cell(r, base + 1 + i, s3(roles[rc]))
        for i, gc in enumerate(give_cols):
            if s3(gives.get(gc)): ws.cell(r, base + len(role_cols) + 1 + i, s3(gives[gc]))
        for i, nc in enumerate(need_cols):
            if s3(needs.get(nc)): ws.cell(r, base + len(role_cols) + len(give_cols) + 1 + i, s3(needs[nc]))
        if note:
            ws.cell(r, notes_col, note).alignment = left
        for col in range(1, ncol + 1):
            cell = ws.cell(r, col)
            cell.border = BORDER
            if 3 < col < notes_col:
                cell.alignment = center
                cell.font = Font(color="1D9E75" if col <= 3 + len(role_cols) + len(give_cols) and col > 3 + len(role_cols) else ("534AB7" if col <= 3 + len(role_cols) else "BA7517"))
        r += 1

    # widths + freeze — every tag/role column wide enough for its full header
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    for col in range(4, ncol):
        name = ws.cell(2, col).value or ""
        ws.column_dimensions[get_column_letter(col)].width = max(6, len(name) + 3)
    ws.column_dimensions[get_column_letter(ncol)].width = 50  # заметки
    ws.freeze_panes = "D3"
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20

    # dictionary sheet
    ds = wb.create_sheet("Словарь")
    ds.append(["РОЛИ (можно несколько)"])
    ds["A1"].font = Font(bold=True)
    for k, v in ROLES:
        ds.append([k, v])
    ds.append([])
    ds.append(["ТЕГИ (один список — работает и в ДАЁТ, и в НУЖНО)"])
    ds.cell(ds.max_row, 1).font = Font(bold=True)
    for k, v in TAGS:
        ds.append([k, v])
    ds.append([])
    ds.append(["ШКАЛА: пусто=нет · 1 · 2 · 3 · 4 (5 уровней). Роли — насколько исполняет роль."])
    ds.append(["ДАЁТ — сила эффекта. НУЖНО — насколько критично агенту (децибелы обычно 1)."])
    ds.append(["Роль ставим по СООТВЕТСТВИЮ (мейн крит-дпс, не офф-филд), не по качеству — качество ловит индивид-WR."])
    ds.append(["Qingyi ДАЁТ amp_on_stun; Шуньгуан его НУЖНО. Тег один, колонки разные."])
    ds.append(["Ширики (sheer) скейлятся от HP → atk_buff им НУЖНО слабо (~1)."])
    ds.append(["БАЗА = только кит агента. Сигны/пушки/сеты — ситуативны: в число не вписываем (1 = ситуативно), детали в заметки."])
    ds.column_dimensions["A"].width = 20
    ds.column_dimensions["B"].width = 60
    for row in ds.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    out = os.path.join(HERE, "synergy_tags_v3.xlsx")
    wb.save(out)
    print("wrote", out, "with", r - 3, "agents")


if __name__ == "__main__":
    build()
